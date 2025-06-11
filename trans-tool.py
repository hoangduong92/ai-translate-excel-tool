import os
from openai import OpenAI
from openpyxl import load_workbook
import time
from tqdm import tqdm

# Global constants
API_DELAY = 2  # Default delay in seconds
# System prompt filename - change this to switch between different translation directions
# Options: "system_prompt_vi_to_ja.txt", "system_prompt_ja_to_vi.txt", etc.
SYSTEM_PROMPT_FILE = "system_prompt_ja_to_vi.txt"

# Target language mapping based on system prompt filename
LANGUAGE_MAPPING = {
    "system_prompt_vi_to_ja.txt": "Japanese",
    "system_prompt_ja_to_vi.txt": "Vietnamese",
    # Add more language pairs as needed
}
import argparse
import threading
import logging
from datetime import datetime
import signal
import sys
import json
import re
from langdetect import detect
from concurrent.futures import ThreadPoolExecutor, as_completed
import traceback  # Thêm import này nếu bạn sử dụng traceback
from dotenv import load_dotenv
import math
load_dotenv()

# Tham số global
BATCH_SIZE = 20  # Số dòng Excel gom lại để dịch trong một lần gọi API
API_DELAY = 2.0  # Thời gian chờ giữa các lần gọi API (giây)
GROUP_BY_ROW = True  # Gom các ô theo dòng thay vì riêng lẻ

print(os.getenv("GEMINI_API_KEY"))

client = OpenAI(
    # base_url="https://api.groq.com/openai/v1",
    #base_url="https://api.openai.com/v1/",
    base_url="https://generativelanguage.googleapis.com/v1beta/openai/",
    api_key=os.getenv("GEMINI_API_KEY"),
   # organization=os.getenv("GROQ_AI_ORG")
)

class ExcelTranslator:
    def __init__(self, workers=3, cache_file=None, log_file=None, api_delay=API_DELAY):
        """
        Khởi tạo translator với OpenAI API
        """
        # Thiết lập logging trước
        self.setup_logging(log_file)

        # Lấy API key từ biến môi trường
        api_key = os.getenv('GEMINI_API_KEY')
        if not api_key:
            self.logger.error(
                "Không tìm thấy GEMINI_API_KEY trong biến môi trường")
            raise ValueError(
                "Không tìm thấy GEMINI_API_KEY trong biến môi trường")

        # Khởi tạo các thuộc tính khác
        client.api_key = api_key
        # self.cached_translations = {}
        self.workers = workers
        self.should_exit = False
        # self.cache_file = cache_file
        self.max_retries = 3
        self.base_delay = api_delay  # Sử dụng tham số api_delay

        # Khóa để đảm bảo thread-safety
        # self.cache_lock = threading.Lock()

        # Load cache và thiết lập signal handler
        # self.load_cache()
        signal.signal(signal.SIGINT, self.handle_sigint)

    def setup_logging(self, log_file=None):
        """Thiết lập logging với định dạng chi tiết hơn"""
        if log_file is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            log_file = f"translation_log_{timestamp}.log"

        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - [%(threadName)s] - %(message)s',
            handlers=[
                logging.FileHandler(log_file, encoding='utf-8'),
                logging.StreamHandler(sys.stdout)
            ]
        )
        self.logger = logging.getLogger(__name__)

    def load_cache(self):
        """Nạp cache từ file JSON"""
        # if self.cache_file and os.path.exists(self.cache_file):
        #     try:
        #         with open(self.cache_file, 'r', encoding='utf-8') as f:
        #             self.cached_translations = json.load(f)
        #         self.logger.info(
        #             f"Đã nạp {len(self.cached_translations)} bản dịch từ cache")
        #     except Exception as e:
        #         self.logger.error(f"Lỗi khi nạp cache: {str(e)}")
        pass

    def save_cache(self):
        """Lưu cache vào file JSON"""
        # if self.cache_file:
        #     with self.cache_lock:
        #         try:
        #             with open(self.cache_file, 'w', encoding='utf-8') as f:
        #                 json.dump(self.cached_translations, f,
        #                           ensure_ascii=False, indent=2)
        #             self.logger.info(
        #                 f"Đã lưu {len(self.cached_translations)} bản dịch vào cache")
        #         except Exception as e:
        #             self.logger.error(f"Lỗi khi lưu cache: {str(e)}")
        pass

    def handle_sigint(self, signum, frame):
        """Xử lý tín hiệu Ctrl+C"""
        self.logger.info("\nĐang dừng chương trình...")
        self.should_exit = True
        self.save_cache()
        sys.exit(0)

    def clean_text(self, text):
        """
        Chuẩn hóa text trước khi dịch
        """
        if not isinstance(text, str):
            return str(text)

        text = ''.join(char for char in text if ord(
            char) >= 32 or char in '\n\t')
        text = ' '.join(text.split())

        return text.strip()

    def should_translate_cell(self, cell, cell_info=""):
        """
        Kiểm tra xem một ô có cần dịch hay không
        """
        if cell.value is None:
            return False, "Ô trống"

        value = self.clean_text(str(cell.value))

        if not value:
            return False, "Ô chỉ chứa khoảng trắng hoặc ký tự đặc biệt"

        # Kiểm tra nếu là ký tự đơn lẻ
        if len(value) <= 1:
            # Kiểm tra xem có phải là ký tự kanji hoặc hiragana/katakana hay không
            # hoặc bất kỳ ký tự quan trọng khác
            if any(ord(c) > 0x4E00 for c in value):  # Kiểm tra phạm vi của các ký tự kanji
                return True, "Ký tự kanji cần dịch"
            return False, "Văn bản quá ngắn"

        if re.match(r'^[\d\s,.-]+$', value):
            return False, "Ô chỉ chứa số và ký tự định dạng số"

        if value.startswith('='):
            return False, "Ô chứa công thức Excel"

        # # Phát hiện ngôn ngữ
        # try:
        #     lang = detect(value)
        #     if lang == 'ja':
        #         return False, "Văn bản đã là tiếng Nhật"
        # except:
        #     pass  # Nếu không phát hiện được ngôn ngữ, tiếp tục xử lý

        return True, "Cần dịch"

    def translate_batch_to_japanese(self, texts, cell_infos=None):
        """
        Dịch một batch văn bản sang tiếng Nhật sử dụng OpenAI API
        
        Args:
            texts: Danh sách các văn bản cần dịch
            cell_infos: Danh sách thông tin về các ô tương ứng
            
        Returns:
            Danh sách các văn bản đã dịch
        """
        if not texts:
            return []
            
        if cell_infos is None:
            cell_infos = ["" for _ in texts]
            
        # Làm sạch văn bản
        cleaned_texts = [self.clean_text(text) for text in texts]
        # Loại bỏ các văn bản trống
        valid_indices = [i for i, text in enumerate(cleaned_texts) if text]
        if not valid_indices:
            return texts  # Trả về danh sách ban đầu nếu không có văn bản hợp lệ
            
        valid_texts = [cleaned_texts[i] for i in valid_indices]
        valid_cell_infos = [cell_infos[i] for i in valid_indices]
        
        # Chuẩn bị kết quả với giá trị ban đầu
        results = list(texts)  # Tạo bản sao của danh sách ban đầu
        
        # Ghép các văn bản thành một chuỗi duy nhất với dấu phân cách
        batch_text = "\n---ITEM_SEPARATOR---\n".join(valid_texts)
        
        for attempt in range(self.max_retries):
            try:
                start_time = time.time()
                
                # Xác định ngôn ngữ đích dựa trên file system prompt
                target_language = LANGUAGE_MAPPING.get(os.path.basename(SYSTEM_PROMPT_FILE), "Japanese")  # Fallback to Japanese if not found
                
                # Tạo prompt yêu cầu dịch batch văn bản
                prompt = f"Translate each of the following text items to {target_language}. Each item is separated by '---ITEM_SEPARATOR---'.\n\n{batch_text}"
                
                # Read system prompt from file
                system_prompt_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), SYSTEM_PROMPT_FILE)
                try:
                    with open(system_prompt_path, 'r', encoding='utf-8') as f:
                        system_prompt = f.read()
                    self.logger.info(f"Successfully loaded system prompt from {system_prompt_path}")
                except Exception as e:
                    self.logger.error(f"Error loading system prompt from file: {e}")
                    # Fallback to default prompt in case of error
                    system_prompt = """You are a professional translator. Follow these rules strictly:
1. Output ONLY the Japanese translations, nothing else
2. DO NOT include the original text in your response
3. DO NOT add any explanations or notes
4. Keep IDs and special characters unchanged
5. Convert regular numbers to Japanese numbers (1->１, 2->２, etc.)
6. Use standard Japanese IT terminology for technical terms
7. Preserve the original formatting (spaces, line breaks)
8. For mixed language text, translate all non-Japanese parts to Japanese
9. Use proper Japanese particle usage (の, を, に, etc.)
10. IMPORTANT: Maintain the same number of items and keep them separated by '---ITEM_SEPARATOR---'"""
                
                response = client.chat.completions.create(
                    model="gemini-2.0-flash-lite",
                    messages=[
                        {"role": "system", "content": system_prompt},
                        {"role": "user", "content": prompt}
                    ]
                )

                translation_time = time.time() - start_time
                translated_text = response.choices[0].message.content
                
                if translated_text:
                    # Tách các bản dịch từ kết quả
                    translated_items = translated_text.split("---ITEM_SEPARATOR---")
                    
                    # Kiểm tra số lượng bản dịch có khớp với số lượng văn bản hợp lệ không
                    if len(translated_items) >= len(valid_texts):
                        # Ghi log thông tin về batch
                        self.logger.info(
                            f"Đã dịch batch với {len(valid_texts)} mục trong {translation_time:.2f}s"
                        )
                        
                        # Ghi log chi tiết cho từng mục
                        for i, (orig, trans, info) in enumerate(zip(valid_texts, translated_items, valid_cell_infos)):
                            self.logger.info(
                                f"{info}\n"
                                f"Văn bản gốc ({len(orig)} ký tự): {orig}\n"
                                f"Bản dịch ({len(trans.strip())} ký tự): {trans.strip()}"
                            )
                        
                        # Cập nhật kết quả
                        for i, idx in enumerate(valid_indices):
                            if i < len(translated_items):
                                results[idx] = translated_items[i].strip()
                        
                        time.sleep(self.base_delay)  # Chờ giữa các lần gọi API
                        return results
                    else:
                        self.logger.warning(
                            f"Số lượng bản dịch không khớp: nhận {len(translated_items)}, mong đợi {len(valid_texts)}"
                        )
                
                self.logger.warning(f"Lần thử {attempt + 1}: Không nhận được kết quả dịch hợp lệ")
                time.sleep(self.base_delay * (attempt + 1))
                
            except Exception as e:
                self.logger.warning(f"Lần thử {attempt + 1} thất bại: {str(e)}")
                if attempt < self.max_retries - 1:
                    time.sleep(self.base_delay * (attempt + 1))
                    continue
                raise
        
        self.logger.error(f"Không thể dịch batch sau {self.max_retries} lần thử")
        return results  # Trả về danh sách ban đầu nếu không dịch được
    
    def translate_to_japanese(self, text, cell_info=""):
        """
        Dịch một văn bản sang tiếng Nhật (wrapper cho hàm batch)
        """
        try:
            text = self.clean_text(text)
            if not text:
                return text
                
            # Gọi hàm dịch batch với một mục duy nhất
            results = self.translate_batch_to_japanese([text], [cell_info])
            return results[0] if results else text
            
        except Exception as e:
            self.logger.error(f"{cell_info} - Lỗi dịch: {str(e)}")
            return text

    def process_excel_file(self, input_path, output_dir=None):
        """
        Xử lý một file Excel: đọc nội dung, xác định các ô cần dịch, thực hiện dịch và lưu kết quả.
        
        Args:
            input_path: Đường dẫn đến file Excel cần dịch
            output_dir: Thư mục đầu ra. Nếu None, lưu file với hậu tố '_translated' cùng thư mục với file gốc
        """
        try:
            base_filename = os.path.basename(input_path)
            filename, ext = os.path.splitext(base_filename)
            
            if output_dir:
                # Đảm bảo thư mục đầu ra tồn tại
                os.makedirs(output_dir, exist_ok=True)
                output_path = os.path.join(output_dir, f"{filename}_translated{ext}")
            else:
                # Nếu không có output_dir, lưu cùng thư mục với file gốc
                filename_full, ext = os.path.splitext(input_path)
                output_path = f"{filename_full}_translated{ext}"

            self.logger.info(
                f"Bắt đầu xử lý file: {os.path.basename(input_path)}")
            wb = load_workbook(input_path, data_only=False)

            file_start_time = time.time()
            translation_stats = {
                'total_cells': 0,
                'processed_cells': 0,
                'translated_cells': 0,
                # 'cached_translations': 0,
                'failed_translations': 0
            }

            for sheet_name in wb.sheetnames:
                if self.should_exit:
                    raise KeyboardInterrupt

                ws = wb[sheet_name]
                sheet_start_time = time.time()

                self.logger.info(f"Bắt đầu xử lý sheet: {sheet_name}")

                # Duyệt qua các ô trong sheet
                cells_to_translate = []

                for row in ws.iter_rows():
                    for cell in row:
                        translation_stats['total_cells'] += 1

                        cell_info = (
                            f"Sheet: {sheet_name}, "
                            f"Ô: {cell.coordinate}"
                        )

                        should_translate, reason = self.should_translate_cell(
                            cell, cell_info)

                        if should_translate:
                            cells_to_translate.append((cell, cell_info))
                            translation_stats['processed_cells'] += 1
                            self.logger.debug(f"{cell_info} - {reason}")
                        else:
                            self.logger.debug(
                                f"{cell_info} - Bỏ qua: {reason}")

                self.logger.info(
                    f"Thống kê sheet {sheet_name}:\n"
                    f"- Tổng số ô: {translation_stats['total_cells']}\n"
                    f"- Ô cần dịch: {translation_stats['processed_cells']}"
                )

                # Dịch các ô theo batch
                with tqdm(total=len(cells_to_translate), desc=f"Dịch sheet {sheet_name}", unit='cell') as pbar:
                    batches = []
                    
                    if GROUP_BY_ROW:
                        # Gom các ô theo dòng
                        cells_by_row = {}
                        for cell, cell_info in cells_to_translate:
                            # Lấy số dòng từ ô
                            row_num = cell.row
                            if row_num not in cells_by_row:
                                cells_by_row[row_num] = []
                            cells_by_row[row_num].append((cell, cell_info))
                        
                        # Chuyển từ dict sang list các dòng
                        rows = [cells_by_row[row_num] for row_num in sorted(cells_by_row.keys())]
                        
                        # Chia các dòng thành các batch có kích thước BATCH_SIZE
                        row_batches = [rows[i:i + BATCH_SIZE] for i in range(0, len(rows), BATCH_SIZE)]
                        
                        # Chuyển từ batch các dòng sang batch các ô
                        for row_batch in row_batches:
                            # Gom tất cả các ô từ các dòng trong batch này
                            cells_batch = []
                            for row in row_batch:
                                cells_batch.extend(row)
                            batches.append(cells_batch)
                            
                        self.logger.info(f"Chia thành {len(batches)} batch theo dòng, mỗi batch tối đa {BATCH_SIZE} dòng")
                    else:
                        # Chia các ô thành các batch có kích thước BATCH_SIZE (cách cũ)
                        batches = [cells_to_translate[i:i + BATCH_SIZE] for i in range(0, len(cells_to_translate), BATCH_SIZE)]
                        self.logger.info(f"Chia thành {len(batches)} batch, mỗi batch tối đa {BATCH_SIZE} ô")
                    
                    for batch_idx, batch in enumerate(batches):
                        if self.should_exit:
                            raise KeyboardInterrupt
                            
                        if GROUP_BY_ROW:
                            self.logger.info(f"Xử lý batch {batch_idx + 1}/{len(batches)} với {len(batch)} ô từ nhiều dòng")
                        else:
                            self.logger.info(f"Xử lý batch {batch_idx + 1}/{len(batches)} với {len(batch)} ô")
                        
                        # Chuẩn bị dữ liệu cho batch
                        batch_texts = [cell.value for cell, _ in batch]
                        batch_cell_infos = [cell_info for _, cell_info in batch]
                        
                        try:
                            # Dịch cả batch
                            translated_texts = self.translate_batch_to_japanese(batch_texts, batch_cell_infos)
                            
                            # Cập nhật giá trị cho các ô
                            for i, ((cell, _), translated_text) in enumerate(zip(batch, translated_texts)):
                                if translated_text != cell.value:
                                    cell.value = translated_text
                                    translation_stats['translated_cells'] += 1
                                pbar.update(1)
                                
                        except Exception as e:
                            self.logger.error(f"Lỗi khi xử lý batch {batch_idx + 1}: {str(e)}")
                            # Nếu xử lý batch thất bại, thử xử lý từng ô một
                            self.logger.info("Thử xử lý từng ô riêng lẻ...")
                            
                            with ThreadPoolExecutor(max_workers=self.workers) as executor:
                                future_to_cell = {
                                    executor.submit(
                                        self.translate_to_japanese,
                                        cell.value,
                                        cell_info
                                    ): (cell, cell_info)
                                    for cell, cell_info in batch
                                }
                                
                                for future in as_completed(future_to_cell):
                                    if self.should_exit:
                                        raise KeyboardInterrupt
                                        
                                    cell, cell_info = future_to_cell[future]
                                    try:
                                        translated_text = future.result()
                                        if translated_text != cell.value:
                                            cell.value = translated_text
                                            translation_stats['translated_cells'] += 1
                                    except Exception as e:
                                        self.logger.error(f"{cell_info} - Lỗi dịch: {str(e)}")
                                        translation_stats['failed_translations'] += 1
                                    finally:
                                        pbar.update(1)

                sheet_time = time.time() - sheet_start_time
                self.logger.info(
                    f"Hoàn thành sheet {sheet_name}:\n"
                    f"- Thời gian xử lý: {sheet_time:.2f}s"
                )

            wb.save(output_path)
            self.logger.info(
                f"Đã lưu file dịch vào: {output_path}\n"
                f"- Tổng số ô: {translation_stats['total_cells']}\n"
                f"- Ô cần dịch: {translation_stats['processed_cells']}\n"
                f"- Ô đã dịch: {translation_stats['translated_cells']}\n"
                f"- Ô dịch thất bại: {translation_stats['failed_translations']}"
            )

        except KeyboardInterrupt:
            self.logger.info("Đang dừng xử lý file...")
            try:
                wb.save(output_path)
                self.logger.info(f"Đã lưu các thay đổi vào: {output_path}")
            except:
                self.logger.error("Không thể lưu file.")
            raise
        except Exception as e:
            self.logger.error(f"Lỗi khi xử lý file: {str(e)}")
            raise


def main():
    # Khai báo các biến global sẽ được sử dụng trong hàm này
    global BATCH_SIZE, API_DELAY, GROUP_BY_ROW
    
    # Xác định đường dẫn đến thư mục chứa script
    script_dir = os.path.dirname(os.path.abspath(__file__))
    
    # Đường dẫn mặc định cho thư mục input và output
    default_input_dir = os.path.join(script_dir, 'input')
    default_output_dir = os.path.join(script_dir, 'output')
    
    parser = argparse.ArgumentParser(
        description='Dịch nội dung file Excel sang tiếng Nhật')
    parser.add_argument('-i', '--input',
                        help='Đường dẫn đến thư mục chứa file Excel cần dịch (mặc định: thư mục "input" cùng cấp với script)',
                        default=default_input_dir)
    parser.add_argument('-o', '--output',
                        help='Đường dẫn đến thư mục lưu file Excel đã dịch (mặc định: thư mục "output" cùng cấp với script)',
                        default=default_output_dir)
    parser.add_argument('-d', '--directory',
                        help='Đường dẫn đến thư mục chứa file Excel cần dịch (tương thích ngược)',
                        default=None)
    parser.add_argument('-f', '--file',
                        help='Đường dẫn đến file Excel cần dịch',
                        default=None)
    parser.add_argument('-w', '--workers',
                        help='Số lượng worker threads (mặc định: 3)',
                        type=int,
                        default=3)
    parser.add_argument('-b', '--batch-size',
                        help='Số lượng dòng Excel gom lại để dịch trong một lần gọi API (mặc định: 20)',
                        type=int,
                        default=BATCH_SIZE)
    parser.add_argument('--group-by-row',
                        help='Gom các ô theo dòng thay vì riêng lẻ (mặc định: True)',
                        action='store_true',
                        default=GROUP_BY_ROW)
    parser.add_argument('-a', '--api-delay',
                        help='Thời gian chờ giữa các lần gọi API (giây) (mặc định: 1.0)',
                        type=float,
                        default=API_DELAY)
    # parser.add_argument('-c', '--cache',
    #                     help='File cache để lưu các bản dịch',
    #                     default='translation_cache.json')
    parser.add_argument('-l', '--log',
                        help='File log output',
                        default=None)

    args = parser.parse_args()

    try:
        # Cập nhật tham số global nếu được chỉ định qua command line
        if args.batch_size != BATCH_SIZE:
            BATCH_SIZE = args.batch_size
            print(f"Đã đặt kích thước batch: {BATCH_SIZE}")
            
        if args.api_delay != API_DELAY:
            API_DELAY = args.api_delay
            print(f"Đã đặt thời gian chờ API: {API_DELAY}s")
            
        if args.group_by_row != GROUP_BY_ROW:
            GROUP_BY_ROW = args.group_by_row
            print(f"Chế độ gom theo dòng: {'Bật' if GROUP_BY_ROW else 'Tắt'}")
            
        print(f"Khởi tạo translator...")
        translator = ExcelTranslator(
            workers=args.workers,
            # cache_file=args.cache,
            log_file=args.log,
            api_delay=API_DELAY
        )

        # Xác định thư mục đầu vào (ưu tiên --directory nếu được cung cấp, để tương thích ngược)
        input_dir = args.directory if args.directory else args.input
        output_dir = args.output
        
        print(f"Thư mục đầu vào: {input_dir}")
        print(f"Thư mục đầu ra: {output_dir}")
        
        # Process a single file if specified
        if args.file:
            if not os.path.isfile(args.file):
                print(f"Lỗi: File '{args.file}' không tồn tại!")
                return
                
            if not args.file.endswith('.xlsx') or args.file.endswith('_translated.xlsx'):
                print(f"Lỗi: '{args.file}' không phải là file Excel hợp lệ!")
                return
                
            print(f"\nĐang xử lý file: {os.path.basename(args.file)}")
            translator.process_excel_file(args.file, output_dir)
            return

        print(f"Kiểm tra thư mục đầu vào: {input_dir}")
        if not os.path.exists(input_dir):
            print(f"Lỗi: Thư mục '{input_dir}' không tồn tại!")
            return
            
        if not os.path.isdir(input_dir):
            print(f"Lỗi: '{input_dir}' không phải là thư mục!")
            return
            
        # Kiểm tra thư mục đầu ra nếu được cung cấp
        if output_dir:
            if os.path.exists(output_dir) and not os.path.isdir(output_dir):
                print(f"Lỗi: '{output_dir}' không phải là thư mục!")
                return
            # Tạo thư mục đầu ra nếu chưa tồn tại
            os.makedirs(output_dir, exist_ok=True)

        print("Tìm các file Excel...")
        excel_files = [
            f for f in os.listdir(input_dir)
            if f.endswith('.xlsx') and not f.endswith('_translated.xlsx')
        ]

        if not excel_files:
            print(
                f"Không tìm thấy file Excel nào trong thư mục '{input_dir}'")
            return

        print(f"Tìm thấy {len(excel_files)} file Excel:")
        for filename in excel_files:
            print(f"- {filename}")

        print("\nBắt đầu xử lý các file:")
        for filename in excel_files:
            file_path = os.path.join(input_dir, filename)
            try:
                print(f"\nĐang xử lý file: {filename}")
                translator.process_excel_file(file_path, output_dir)
            except KeyboardInterrupt:
                print("\nĐã dừng chương trình theo yêu cầu.")
                break
            except Exception as e:
                print(f"Lỗi khi xử lý file {filename}: {str(e)}")

    except KeyboardInterrupt:
        print("\nĐã dừng chương trình.")
        sys.exit(0)
    except Exception as e:
        print(f"Lỗi: {str(e)}")
        traceback.print_exc()  # In ra stack trace đầy đủ
        sys.exit(1)


if __name__ == "__main__":
    main()
